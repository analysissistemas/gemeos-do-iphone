"""
Gera a planilha do catálogo da loja.

Entrada:  _catalogo.json  (exportado pelo próprio sistema, para a planilha
          nunca discordar do que está na tela)
Saída:    Catalogo-Gemeos-do-iPhone.xlsx  na Área de Trabalho

Três abas:
  Catálogo  — uma linha por produto, com custo, venda, lucro e margem
  Resumo    — totais por categoria
  Ajuda     — o que cada coluna quer dizer

Os valores de lucro e margem são FÓRMULAS, não números colados: se você mudar
o custo ou o preço na planilha, o resto se recalcula sozinho.
"""
import json
import os
from collections import OrderedDict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

AQUI = os.path.dirname(os.path.abspath(__file__))
ENTRADA = os.path.join(AQUI, "_catalogo.json")

desktop = os.path.join(os.environ["USERPROFILE"], "Desktop")
if not os.path.isdir(desktop):
    desktop = os.path.join(os.environ["USERPROFILE"], "OneDrive", "Desktop")
SAIDA = os.path.join(desktop, "Catalogo-Gemeos-do-iPhone.xlsx")

# ---- identidade da loja ----
PRETO = "0B0B0C"
CINZA_CAB = "1F2937"
LINHA_PAR = "F7F9FC"
BORDA = "D8DEE9"
VERDE = "0A7D2E"
VERMELHO = "B42318"
FONTE = "Arial"

with open(ENTRADA, encoding="utf-8") as f:
    dados = json.load(f)
produtos = dados["produtos"]

COLUNAS = [
    ("ID", 6, "id"),
    ("Categoria", 16, "categoria"),
    ("Modelo", 24, "modelo"),
    ("Memória / Tamanho", 17, "variacao"),
    ("Cor", 20, "cor"),
    ("Condição", 15, "condicao"),
    ("Bateria %", 10, "bateria"),
    ("Avarias", 26, "avarias"),
    ("IMEI", 17, "imei"),
    ("Entrada", 11, "entrada"),
    ("Custo", 12, "custo"),
    ("Venda", 12, "venda"),
    ("Lucro", 12, None),         # fórmula
    ("Margem", 10, None),        # fórmula
    ("Qtd", 7, "quantidade"),
    ("Controle", 13, "controle"),
    ("Situação", 13, "situacao"),
    ("Tem foto", 10, "temFoto"),
    ("Na vitrine", 11, "naVitrine"),
]

fina = Side(style="thin", color=BORDA)
borda = Border(left=fina, right=fina, top=fina, bottom=fina)

wb = Workbook()

# ============================================================ ABA CATÁLOGO
ws = wb.active
ws.title = "Catálogo"

ws["A1"] = "Catálogo — Gêmeos do iPhone"
ws["A1"].font = Font(name=FONTE, size=15, bold=True, color=PRETO)
ws["A2"] = f"Exportado do sistema em {dados['gerado'][:10]} · {len(produtos)} produtos"
ws["A2"].font = Font(name=FONTE, size=9, color="6B7280")
ws.merge_cells("A1:E1")
ws.merge_cells("A2:E2")

CAB = 4
for i, (titulo, larg, _) in enumerate(COLUNAS, start=1):
    c = ws.cell(row=CAB, column=i, value=titulo)
    c.font = Font(name=FONTE, size=10, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=CINZA_CAB)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = borda
    ws.column_dimensions[get_column_letter(i)].width = larg
ws.row_dimensions[CAB].height = 28

for n, p in enumerate(produtos):
    r = CAB + 1 + n
    for i, (titulo, _, chave) in enumerate(COLUNAS, start=1):
        if chave is None:
            continue
        ws.cell(row=r, column=i, value=p.get(chave, ""))

    # Lucro e Margem como FÓRMULA: mexeu no custo ou no preço, recalcula sozinho.
    # A margem é protegida contra venda zerada, senão daria erro de divisão.
    ws.cell(row=r, column=13, value=f"=L{r}-K{r}")
    ws.cell(row=r, column=14, value=f'=IFERROR((L{r}-K{r})/L{r},"")')

    for i in range(1, len(COLUNAS) + 1):
        c = ws.cell(row=r, column=i)
        c.font = Font(name=FONTE, size=10)
        c.border = borda
        if n % 2 == 1:
            c.fill = PatternFill("solid", fgColor=LINHA_PAR)
    for col in (11, 12, 13):
        ws.cell(row=r, column=col).number_format = 'R$ #,##0;-R$ #,##0;"-"'
    ws.cell(row=r, column=14).number_format = "0.0%"
    ws.cell(row=r, column=13).font = Font(name=FONTE, size=10, bold=True, color=VERDE)
    for col in (1, 7, 15):
        ws.cell(row=r, column=col).alignment = Alignment(horizontal="center")
    ws.cell(row=r, column=9).font = Font(name=FONTE, size=9, color="6B7280")

ULTIMA = CAB + len(produtos)

# linha de totais
t = ULTIMA + 1
ws.cell(row=t, column=1, value="TOTAL").font = Font(name=FONTE, size=10, bold=True)
ws.merge_cells(start_row=t, start_column=1, end_row=t, end_column=10)
for col, letra in ((11, "K"), (12, "L"), (13, "M")):
    c = ws.cell(row=t, column=col, value=f"=SUM({letra}{CAB+1}:{letra}{ULTIMA})")
    c.font = Font(name=FONTE, size=10, bold=True)
    c.number_format = 'R$ #,##0;-R$ #,##0;"-"'
    c.border = Border(top=Side(style="medium", color=CINZA_CAB))
ws.cell(row=t, column=14, value=f'=IFERROR(M{t}/L{t},"")')
ws.cell(row=t, column=14).number_format = "0.0%"
ws.cell(row=t, column=14).font = Font(name=FONTE, size=10, bold=True)
ws.cell(row=t, column=1).border = Border(top=Side(style="medium", color=CINZA_CAB))

ws.freeze_panes = f"A{CAB+1}"
ws.auto_filter.ref = f"A{CAB}:{get_column_letter(len(COLUNAS))}{ULTIMA}"

# ============================================================ ABA RESUMO
rs = wb.create_sheet("Resumo")
rs["A1"] = "Resumo por categoria"
rs["A1"].font = Font(name=FONTE, size=15, bold=True, color=PRETO)
rs["A2"] = "Os números vêm da aba Catálogo — mexeu lá, muda aqui."
rs["A2"].font = Font(name=FONTE, size=9, color="6B7280")

cats = list(OrderedDict.fromkeys(p["categoria"] for p in produtos))
cabs = ["Categoria", "Produtos", "Investido (custo)", "A receber (venda)", "Lucro previsto", "Margem"]
larguras = [22, 12, 19, 19, 17, 11]
for i, (h, w) in enumerate(zip(cabs, larguras), start=1):
    c = rs.cell(row=4, column=i, value=h)
    c.font = Font(name=FONTE, size=10, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=CINZA_CAB)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = borda
    rs.column_dimensions[get_column_letter(i)].width = w
rs.row_dimensions[4].height = 26

faixa = f"Catálogo!$B${CAB+1}:$B${ULTIMA}"
for n, cat in enumerate(cats):
    r = 5 + n
    rs.cell(row=r, column=1, value=cat)
    rs.cell(row=r, column=2, value=f'=COUNTIF({faixa},$A{r})')
    rs.cell(row=r, column=3, value=f'=SUMIF({faixa},$A{r},Catálogo!$K${CAB+1}:$K${ULTIMA})')
    rs.cell(row=r, column=4, value=f'=SUMIF({faixa},$A{r},Catálogo!$L${CAB+1}:$L${ULTIMA})')
    rs.cell(row=r, column=5, value=f"=D{r}-C{r}")
    rs.cell(row=r, column=6, value=f'=IFERROR(E{r}/D{r},"")')
    for i in range(1, 7):
        c = rs.cell(row=r, column=i)
        c.font = Font(name=FONTE, size=10)
        c.border = borda
        if n % 2 == 1:
            c.fill = PatternFill("solid", fgColor=LINHA_PAR)
    for col in (3, 4, 5):
        rs.cell(row=r, column=col).number_format = 'R$ #,##0;-R$ #,##0;"-"'
    rs.cell(row=r, column=6).number_format = "0.0%"
    rs.cell(row=r, column=2).alignment = Alignment(horizontal="center")
    rs.cell(row=r, column=5).font = Font(name=FONTE, size=10, bold=True, color=VERDE)

tr = 5 + len(cats)
rs.cell(row=tr, column=1, value="TOTAL").font = Font(name=FONTE, size=10, bold=True)
for col, letra in ((2, "B"), (3, "C"), (4, "D"), (5, "E")):
    c = rs.cell(row=tr, column=col, value=f"=SUM({letra}5:{letra}{tr-1})")
    c.font = Font(name=FONTE, size=10, bold=True)
    c.border = Border(top=Side(style="medium", color=CINZA_CAB))
    if col > 2:
        c.number_format = 'R$ #,##0;-R$ #,##0;"-"'
rs.cell(row=tr, column=2).alignment = Alignment(horizontal="center")
rs.cell(row=tr, column=6, value=f'=IFERROR(E{tr}/D{tr},"")')
rs.cell(row=tr, column=6).number_format = "0.0%"
rs.cell(row=tr, column=6).font = Font(name=FONTE, size=10, bold=True)
rs.cell(row=tr, column=1).border = Border(top=Side(style="medium", color=CINZA_CAB))

# ============================================================ ABA AJUDA
aj = wb.create_sheet("Ajuda")
aj["A1"] = "Como ler esta planilha"
aj["A1"].font = Font(name=FONTE, size=15, bold=True, color=PRETO)
aj.column_dimensions["A"].width = 26
aj.column_dimensions["B"].width = 96

EXPLICA = [
    ("Custo", "O que você pagou no aparelho. É o número que você digita — pode corrigir aqui."),
    ("Venda", "Preço pedido ao cliente. Também pode ser corrigido aqui."),
    ("Lucro", "FÓRMULA: Venda menos Custo. Muda sozinho se você corrigir um dos dois."),
    ("Margem", "FÓRMULA: quanto do preço de venda é lucro. 30% quer dizer que R$30 de cada R$100 sobram."),
    ("Bateria %", "Saúde da bateria daquele aparelho. Vazio quer dizer acessório, ou aparelho lacrado."),
    ("Avarias", "Defeitos declarados. Vazio quer dizer sem avaria."),
    ("Controle", "«peça única» é aparelho com IMEI, contado um a um. «quantidade» é acessório, contado por unidade."),
    ("Situação", "Se o item ainda está no estoque ou já foi vendido."),
    ("Tem foto", "Se existe foto do produto no catálogo do cliente. «não» aparece com o contorno desenhado."),
    ("Na vitrine", "Se o cliente enxerga esse item na loja. «não» fica só no seu controle interno."),
]
aj["A3"] = "Coluna"
aj["B3"] = "O que significa"
for c in ("A3", "B3"):
    aj[c].font = Font(name=FONTE, size=10, bold=True, color="FFFFFF")
    aj[c].fill = PatternFill("solid", fgColor=CINZA_CAB)
    aj[c].border = borda
for i, (col, txt) in enumerate(EXPLICA):
    r = 4 + i
    aj.cell(row=r, column=1, value=col).font = Font(name=FONTE, size=10, bold=True)
    aj.cell(row=r, column=2, value=txt).font = Font(name=FONTE, size=10)
    for j in (1, 2):
        aj.cell(row=r, column=j).border = borda
        aj.cell(row=r, column=j).alignment = Alignment(vertical="top", wrap_text=True)
        if i % 2 == 1:
            aj.cell(row=r, column=j).fill = PatternFill("solid", fgColor=LINHA_PAR)
    aj.row_dimensions[r].height = 26

nota = 4 + len(EXPLICA) + 1
aj.cell(row=nota, column=1, value="De onde vêm os dados")
aj.cell(row=nota, column=1).font = Font(name=FONTE, size=11, bold=True)
aj.cell(row=nota + 1, column=1,
        value="Exportado do próprio sistema da loja, para a planilha nunca discordar do que "
              "está na tela. Para atualizar, exporte de novo e rode gerar_excel_catalogo.py.")
aj.cell(row=nota + 1, column=1).font = Font(name=FONTE, size=10, color="6B7280")
aj.merge_cells(start_row=nota + 1, start_column=1, end_row=nota + 1, end_column=2)
aj.cell(row=nota + 3, column=1,
        value="ATENÇÃO: os valores desta versão são dados de exemplo do protótipo, "
              "não o estoque real da loja.")
aj.cell(row=nota + 3, column=1).font = Font(name=FONTE, size=10, bold=True, color=VERMELHO)
aj.merge_cells(start_row=nota + 3, start_column=1, end_row=nota + 3, end_column=2)

wb.save(SAIDA)
print("gerado:", SAIDA)
print("produtos:", len(produtos), "| categorias:", len(cats))
